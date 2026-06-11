"""
GÖREV 1 — Başarı Oranı Dosyasından Ders Birim Aktarımı
=======================================================
DÜZELTİLEN SORUNLAR:
  [1] Tek dönem → 4 dönemin tamamı DONEM_CIFTLERI listesiyle işlenir
  [2] Eşleşmeyen satır raporu → hangi ders kodları eşleşmedi, detaylı çıktı
  [3] Dosya yolları → BASE_KLASOR değişkeni ile taşınabilir hale getirildi

Kullanım:
  • BASE_KLASOR'u xlsx dosyalarının bulunduğu klasöre ayarlayın
  • DONEM_CIFTLERI içindeki dosya adlarını kontrol edin
  • python gorev1_birim_aktar.py
"""

import os
import pandas as pd
from openpyxl import load_workbook


# ============================================================
# TEMEL KLASÖR  →  tüm xlsx dosyalarının bulunduğu dizin
# ============================================================
BASE_KLASOR = "."      # örn: r"C:\Kullanıcı\Belgeler\TBMYO"


# ============================================================
# DÖNEM ÇİFTLERİ  →  (başarı_dosyası, anket_dosyası, çıktı_dosyası, dönem_adı)
# Dosya adlarını kendi bilgisayarınıza göre düzenleyin.
# ============================================================
DONEM_CIFTLERI = [
    {
        "basari" : "2023-2024 Bahar Raporu.xlsx",
        "anket"  : "2023-2024_bahar_anket.xlsx",
        "cikti"  : "2023-2024_bahar_GUNCELLENDI.xlsx",
        "ad"     : "2023-2024 Bahar",
    },
    {
        "basari" : "2024-2025 Güz_basari_oranı_Rapor.xlsx",
        "anket"  : "tbmyo_2024-2025_guz.xlsx",
        "cikti"  : "tbmyo_2024-2025_guz_GUNCELLENDI.xlsx",
        "ad"     : "2024-2025 Güz",
    },
    {
        "basari" : "2024-2025 Bahar_basari_oranı_Rapor.xlsx",
        "anket"  : "tbmyo_2024-2025_bahar.xlsx",
        "cikti"  : "tbmyo_2024-2025_bahar_GUNCELLENDI.xlsx",
        "ad"     : "2024-2025 Bahar",
    },
    {
        "basari" : "2025-2026 Güz_basari_oranı.xlsx",
        "anket"  : "tbmyo_2025-2026_guz.xlsx",
        "cikti"  : "tbmyo_2025-2026_guz_GUNCELLENDI.xlsx",
        "ad"     : "2025-2026 Güz",
    },
]


# ============================================================
# YARDIMCI: anahtar üret  (DERSK0DU_GrupNo, örn. MTD101_1)
# ============================================================
def anahtar_uret(df, kod_sutun, grup_sutun):
    return (
        df[kod_sutun].astype(str).str.strip().str.upper()
        + "_"
        + df[grup_sutun].astype(str)
                        .str.replace(r"\.0$", "", regex=True)
                        .str.strip()
    )


# ============================================================
# TEK DÖNEM İŞLE
# ============================================================
def birim_aktar_tek_donem(basari_yolu, anket_yolu, cikti_yolu, donem_adi):
    print(f"\n{'─' * 55}")
    print(f"  Dönem : {donem_adi}")
    print(f"{'─' * 55}")

    # ----- 1. Dosyaları oku -----
    df_basari = pd.read_excel(basari_yolu)
    df_anket  = pd.read_excel(anket_yolu)

    # Sütun adlarındaki boşlukları ve görünmez karakterleri temizle
    df_basari.columns = df_basari.columns.str.strip()
    df_anket.columns  = df_anket.columns.str.strip()

    print(f"  Başarı : {len(df_basari):>6} satır  →  {os.path.basename(basari_yolu)}")
    print(f"  Anket  : {len(df_anket):>6} satır  →  {os.path.basename(anket_yolu)}")

    # ----- 2. Eşleştirme anahtarı -----
    df_basari["_anahtar"] = anahtar_uret(df_basari, "Ders Kodu", "Grup No")
    df_anket["_anahtar"]  = anahtar_uret(df_anket,  "Ders Kodu", "Grup No")

    # ----- 3. Birim sözlüğü oluştur -----
    birim_map = df_basari.set_index("_anahtar")["Birim"].to_dict()

    # ----- 4. Anket dosyasına yaz -----
    onceki_dolu = df_anket["Ders Birim"].notna().sum()

    df_anket["Ders Birim"] = (
        df_anket["_anahtar"]
        .map(birim_map)
        .fillna(df_anket["Ders Birim"])
    )

    sonraki_dolu = df_anket["Ders Birim"].notna().sum()
    doldurulan   = sonraki_dolu - onceki_dolu

    # ----- 5. [YENİ] Eşleşmeyen ders kodlarını raporla -----
    eslesmeyenler = df_anket[df_anket["Ders Birim"].isna()].copy()
    if not eslesmeyenler.empty:
        eslesmyen_kodlar = (
            eslesmeyenler
            .groupby(["Ders Kodu", "Grup No"])
            .size()
            .reset_index(name="Satır Sayısı")
            .sort_values("Satır Sayısı", ascending=False)
        )
        print(f"\n  ⚠  Eşleşmeyen {len(eslesmeyenler)} satır "
              f"({len(eslesmyen_kodlar)} benzersiz Ders Kodu/Grup):")
        for _, row in eslesmyen_kodlar.iterrows():
            print(f"     • {str(row['Ders Kodu']):15s}  Grup {row['Grup No']}  "
                  f"({row['Satır Sayısı']} satır)")
    else:
        print("\n  ✔  Tüm satırlar eşleşti — eşleşmeyen ders kodu yok.")

    # ----- 6. Orijinal Excel formatını koruyarak kaydet -----
    df_anket.drop(columns=["_anahtar"], inplace=True)

    wb = load_workbook(anket_yolu)
    ws = wb.active

    baslik = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    # Başlıkları temizleyerek ara
    baslik_temiz = [str(b).strip() if b is not None else "" for b in baslik]

    if "Ders Birim" not in baslik_temiz:
        print("  HATA: Anket dosyasında 'Ders Birim' sütunu bulunamadı!")
        return None

    birim_col = baslik_temiz.index("Ders Birim") + 1

    for i, deger in enumerate(df_anket["Ders Birim"], start=2):
        ws.cell(row=i, column=birim_col).value = deger

    wb.save(cikti_yolu)

    # ----- 7. Özet -----
    print(f"\n  Önceden dolu       : {onceki_dolu}")
    print(f"  Yeni doldurulan    : {doldurulan}")
    print(f"  Hâlâ boş           : {len(df_anket) - sonraki_dolu}")
    print(f"\n  ✔ Kaydedildi → '{os.path.basename(cikti_yolu)}'")

    return df_anket


# ============================================================
# ANA FONKSİYON  —  tüm dönemleri işle
# ============================================================
def tum_donemleri_isle():
    print("=" * 55)
    print("  GÖREV 1 — Ders Birim Aktarımı (Tüm Dönemler)")
    print("=" * 55)

    sonuclar = {}
    hatalar  = []

    for d in DONEM_CIFTLERI:
        basari_yolu = os.path.join(BASE_KLASOR, d["basari"])
        anket_yolu  = os.path.join(BASE_KLASOR, d["anket"])
        cikti_yolu  = os.path.join(BASE_KLASOR, d["cikti"])

        # Dosya varlık kontrolü
        if not os.path.exists(basari_yolu):
            print(f"\n  [ATLANDI] Başarı dosyası bulunamadı: {d['basari']}")
            hatalar.append(d["ad"])
            continue
        if not os.path.exists(anket_yolu):
            print(f"\n  [ATLANDI] Anket dosyası bulunamadı: {d['anket']}")
            hatalar.append(d["ad"])
            continue

        try:
            df = birim_aktar_tek_donem(basari_yolu, anket_yolu, cikti_yolu, d["ad"])
            if df is not None:
                sonuclar[d["ad"]] = df
        except Exception as e:
            print(f"\n  HATA ({d['ad']}): {e}")
            hatalar.append(d["ad"])

    # Genel özet
    print(f"\n{'=' * 55}")
    print(f"  Tamamlanan dönemler : {len(sonuclar)}/{len(DONEM_CIFTLERI)}")
    if hatalar:
        print(f"  Başarısız dönemler  : {', '.join(hatalar)}")
    print("=" * 55)

    return sonuclar


if __name__ == "__main__":
    tum_donemleri_isle()
    print("\n✔ Görev 1 tamamlandı.")
