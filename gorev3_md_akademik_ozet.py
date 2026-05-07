"""
GÖREV 3 — MD_Akademik Özet + Ders Detay Sayfaları (Tasarım Bölümü)
====================================================================
Hocamın isterleri:
  1. Her ders için ayrı detay sayfası  →  DersKodu_TIP_BolumAdi
       Sol: Kategori ortalamaları + Katılımcı sayısı
       Sağ: Her sorunun ortalaması + soru metni
  2. Bolum_Genel_Analiz sayfası  →  Kategori bazında genel özet
  3. Tum_Dersler_Ozet sayfası    →  Tüm dersler özet tablosu

  • Mevcut ders sayfaları (_OR_ / _UZ_) korunur, üzerine yazılır
  • Yeni dersler için otomatik sayfa oluşturulur
  • Çıktı: MD_Akademik_OZET.xlsx (orijinali bozulmaz)

DOSYA YOLLARI bölümünü kendi bilgisayarınıza göre güncelleyin.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")


# ============================================================
# DOSYA YOLLARI  →  kendi bilgisayarınıza göre güncelleyin
# ============================================================
TBMYO_GUNCELLENDI = "tbmyo_2025-2026_guz_GUNCELLENDI.xlsx"  # Görev 1 çıktısı
MD_AKADEMIK_DOSYA = "MD_Akademik.xlsx"
CIKTI_DOSYA       = "MD_Akademik_OZET.xlsx"

# Hangi bölüm(ler) alınsın? Boş bırakırsanız tümü alınır.
HEDEF_BOLUM = "Moda Tasarımı"


# ============================================================
# LİKERT DÖNÜŞÜM (metin ve sayı her ikisini de destekler)
# ============================================================
LIKERT_METIN = {
    "kesinlikle katılmıyorum" : 1,
    "katılmıyorum"            : 2,
    "pek fazla katılmıyorum"  : 3,
    "katılıyorum"             : 4,
    "biraz katılıyorum"       : 5,
    "tamamen katılıyorum"     : 6,
}

def likert(v):
    if pd.isna(v): return None
    if isinstance(v, (int, float)):
        x = int(v); return x if 1 <= x <= 6 else None
    if isinstance(v, str):
        return LIKERT_METIN.get(v.strip().lower(), None)
    return None


# ============================================================
# SORU GRUPLARI
# ============================================================
GRUP_ONEKLER = {
    "Ders İçeriği"          : ("1_1", "3_1"),
    "Öğretim Elemanı"       : ("2_1", "4_1", "7_1", "8_1", "9_1", "10_1", "11_1"),
    "Ölçme Değerlendirme"   : ("12_1", "13_1"),
    "Yöntem"                : ("5_1", "6_1"),
    "Genel Memnuniyet"      : ("14_1", "15_1", "16_1"),
}
KAT_LISTESI = list(GRUP_ONEKLER.keys())


# ============================================================
# BİÇİMLENDİRME SABİTLERİ
# ============================================================
R_BASLIK = PatternFill("solid", fgColor="1F4E79")
R_HDR    = PatternFill("solid", fgColor="2E75B6")
R_ALT    = PatternFill("solid", fgColor="D6E4F0")
R_BEYAZ  = PatternFill("solid", fgColor="FFFFFF")
R_TOPLAM = PatternFill("solid", fgColor="FCE4D6")
ince     = Side(style="thin", color="B0B0B0")
KENAR    = Border(left=ince, right=ince, top=ince, bottom=ince)
ORTALA   = Alignment(horizontal="center", vertical="center", wrap_text=True)
SOLA     = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def yaz(cell, val, fill=R_BEYAZ, bold=False, fmt=None, align=ORTALA, fsize=10):
    cell.value     = val
    cell.fill      = fill
    cell.font      = Font(name="Arial", bold=bold, size=fsize,
                          color="FFFFFF" if fill in (R_BASLIK, R_HDR) else "000000")
    cell.alignment = align
    cell.border    = KENAR
    if fmt:
        cell.number_format = fmt


# ============================================================
# YARDIMCI FONKSİYONLAR
# ============================================================
def kategori_etiketi(satir):
    birim   = str(satir.get("Ders Birim", "") or "")
    uzaktan = "uzaktan" in birim.lower()
    kod     = str(satir.get("Ders Kodu", "") or "")
    rakam   = next((c for c in kod if c.isdigit()), "1")
    try:    sinif = min(int(rakam), 2)
    except: sinif = 1
    return f"{sinif}. Sınıf - {'Uzaktan Öğretim' if uzaktan else 'Birinci Öğretim'}"


def sayfa_adi_uret(ders_kodu, birim):
    """DersKodu_TIP_BolumAdi  (TIP: OR=Örgün, UZ=Uzaktan)"""
    tip = "UZ" if "uzaktan" in str(birim).lower() else "OR"
    bolum_temiz = (str(birim)
                   .replace("(Uzaktan Öğretim)", "").replace("(İÖ)", "")
                   .strip())
    raw = f"{ders_kodu}_{tip}_{bolum_temiz}"
    for ch in ("/", "\\", "?", "*", "[", "]", ":"):
        raw = raw.replace(ch, "-")
    return raw[:31]


def temizle_sayfa(ws):
    """Sayfanın tüm hücrelerini sıfırla."""
    # Önce birleşik hücreleri çöz
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))
    # Sonra hücreleri temizle
    for row in ws.iter_rows():
        for cell in row:
            try:
                cell.value = None
                cell.fill  = PatternFill()
                cell.font  = Font()
                cell.border = Border()
                cell.number_format = "General"
            except AttributeError:
                pass


# ============================================================
# 1. DERS DETAY SAYFASI
#    Sol  → kategori ortalamaları
#    Sağ  → soru bazlı ortalamalar (hocamın şablon formatı)
# ============================================================
def ders_detay_yaz(ws, df_ders, ders_kodu, birim, soru_cols):
    ws.sheet_view.showGridLines = False
    n = len(df_ders)

    # Başlık
    ws.merge_cells("A1:B1")
    yaz(ws["A1"], f"{ders_kodu}  |  {birim}",
        fill=R_BASLIK, bold=True, fsize=11, align=SOLA)
    ws.merge_cells("D1:E1")
    yaz(ws["D1"], "Soru Bazlı Ortalamalar",
        fill=R_BASLIK, bold=True, fsize=11)
    ws.row_dimensions[1].height = 28

    # Sol başlıklar
    yaz(ws["A2"], "Kategori",  fill=R_HDR, bold=True)
    yaz(ws["B2"], "Ortalama",  fill=R_HDR, bold=True)
    # Sağ başlıklar
    yaz(ws["D2"], "Soru",      fill=R_HDR, bold=True, align=SOLA)
    yaz(ws["E2"], "Ortalama",  fill=R_HDR, bold=True)
    ws.row_dimensions[2].height = 22

    # Sol: Kategori ortalamaları
    r = 3
    for kat, onekler in GRUP_ONEKLER.items():
        ilgili = [c for c in soru_cols if any(c.startswith(o) for o in onekler)]
        ort    = df_ders[ilgili].mean().mean() if ilgili else float("nan")
        dolgu  = R_ALT if r % 2 == 0 else R_BEYAZ
        yaz(ws.cell(r, 1), kat, dolgu, align=SOLA)
        yaz(ws.cell(r, 2),
            round(ort, 2) if pd.notna(ort) else "-",
            dolgu, fmt="0.00")
        r += 1

    # Memnuniyet ortalaması
    tum = df_ders[soru_cols].mean().mean() if soru_cols else float("nan")
    yaz(ws.cell(r, 1), "Memnuniyet Ortalaması", R_TOPLAM, bold=True, align=SOLA)
    yaz(ws.cell(r, 2), round(tum, 2) if pd.notna(tum) else "-",
        R_TOPLAM, bold=True, fmt="0.00")
    r += 2

    # Katılımcı sayısı
    yaz(ws.cell(r, 1), f"Katılımcı Sayısı (N): {n}",
        R_BEYAZ, bold=True, align=SOLA)
    ws.cell(r, 2).fill = R_BEYAZ

    # Sağ: Soru bazlı ortalamalar
    for i, soru_col in enumerate(soru_cols):
        row_idx = i + 3
        dolgu   = R_ALT if row_idx % 2 == 0 else R_BEYAZ
        ort_s   = df_ders[soru_col].mean()
        yaz(ws.cell(row_idx, 4), soru_col.strip(), dolgu, align=SOLA)
        yaz(ws.cell(row_idx, 5),
            round(ort_s, 2) if pd.notna(ort_s) else "-",
            dolgu, fmt="0.00")
        ws.row_dimensions[row_idx].height = 16

    # Sütun genişlikleri
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 78
    ws.column_dimensions["E"].width = 11
    ws.freeze_panes = "A3"


# ============================================================
# 2. TÜM DERSLER ÖZET
# ============================================================
def tum_dersler_yaz(ws, gruplama):
    ws.sheet_view.showGridLines = False
    SUTUNLAR = [
        "Ders Birim", "Ders Kodu", "Kategori", "Katılımcı (N)",
        "Ders İçeriği", "Öğretim Elemanı", "Ölçme Değerlendirme",
        "Yöntem", "Genel Memnuniyet"
    ]

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value, c.fill = "Tüm Dersler – Ders Değerlendirme Özeti", R_BASLIK
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.alignment = ORTALA; c.border = KENAR
    ws.row_dimensions[1].height = 32

    for ci, ad in enumerate(SUTUNLAR, 1):
        h = ws.cell(2, ci)
        h.value, h.fill = ad, R_HDR
        h.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        h.alignment = ORTALA; h.border = KENAR
    ws.row_dimensions[2].height = 36

    for ri, row in enumerate(gruplama.itertuples(index=False), start=3):
        dolgu = R_ALT if ri % 2 == 0 else R_BEYAZ
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            if ci in (5, 6, 7, 8, 9):
                yaz(cell, round(float(val), 2) if pd.notna(val) else "", dolgu, fmt="0.00")
            elif ci in (1, 3):
                yaz(cell, val, dolgu, align=SOLA)
            else:
                yaz(cell, val, dolgu)
        ws.row_dimensions[ri].height = 18

    son = 2 + len(gruplama)
    if len(gruplama) > 0:
        r = son + 1
        ws.row_dimensions[r].height = 20
        yaz(ws.cell(r, 1), "GENEL ORTALAMA", R_TOPLAM, bold=True, align=SOLA)
        yaz(ws.cell(r, 2), "",  R_TOPLAM)
        yaz(ws.cell(r, 3), "",  R_TOPLAM)
        yaz(ws.cell(r, 4), int(gruplama["Katılımcı (N)"].sum()), R_TOPLAM, bold=True)
        for ci, sc in enumerate(KAT_LISTESI, 5):
            yaz(ws.cell(r, ci), round(gruplama[sc].mean(), 2),
                R_TOPLAM, bold=True, fmt="0.00")

    for ci, g in enumerate([30, 12, 28, 14, 13, 16, 20, 12, 18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = g
    ws.freeze_panes = "A3"


# ============================================================
# 3. BÖLÜM GENEL ANALİZ
# ============================================================
def bolum_genel_yaz(ws, gruplama):
    ws.sheet_view.showGridLines = False
    SUTUNLAR = ["Kategori"] + KAT_LISTESI

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value, c.fill = "Bölüm Genel Analizi – Kategori Ortalamaları", R_BASLIK
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.alignment = ORTALA; c.border = KENAR
    ws.row_dimensions[1].height = 28

    for ci, ad in enumerate(SUTUNLAR, 1):
        h = ws.cell(2, ci)
        h.value, h.fill = ad, R_HDR
        h.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        h.alignment = ORTALA; h.border = KENAR
    ws.row_dimensions[2].height = 36

    kat_ozet = gruplama.groupby("Kategori")[KAT_LISTESI].mean()
    kat_sirali = sorted(kat_ozet.index.tolist())
    kat_ozet = kat_ozet.loc[kat_sirali]

    for ri, (kat, row) in enumerate(kat_ozet.iterrows(), start=3):
        dolgu = R_ALT if ri % 2 == 0 else R_BEYAZ
        yaz(ws.cell(ri, 1), kat, dolgu, align=SOLA)
        for ci, sc in enumerate(KAT_LISTESI, 2):
            yaz(ws.cell(ri, ci),
                round(float(row[sc]), 2) if pd.notna(row[sc]) else "",
                dolgu, fmt="0.00")
        ws.row_dimensions[ri].height = 20

    son = 2 + len(kat_ozet)
    r = son + 1
    ws.row_dimensions[r].height = 20
    yaz(ws.cell(r, 1), "GENEL ORTALAMA", R_TOPLAM, bold=True, align=SOLA)
    for ci, sc in enumerate(KAT_LISTESI, 2):
        yaz(ws.cell(r, ci), round(gruplama[sc].mean(), 2),
            R_TOPLAM, bold=True, fmt="0.00")

    ws.column_dimensions["A"].width = 32
    for ci in range(2, 7):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.freeze_panes = "A3"


# ============================================================
# ANA FONKSİYON
# ============================================================
def md_akademik_ozet_olustur():
    # --- 1. Veri ---
    df = pd.read_excel(TBMYO_GUNCELLENDI)
    print(f"Kaynak: {TBMYO_GUNCELLENDI}  ({len(df)} satır)")

    if HEDEF_BOLUM:
        mask = df["Ders Birim"].astype(str).str.contains(HEDEF_BOLUM, na=False)
        df   = df[mask].copy()
        print(f"Filtre ('{HEDEF_BOLUM}'): {len(df)} satır")

    # --- 2. Likert dönüşümü ---
    soru_cols = [c for c in df.columns if str(c)[0].isdigit()]
    for s in soru_cols:
        df[s] = df[s].apply(likert)

    # --- 3. Kategori etiketi ---
    df["Kategori"] = df.apply(kategori_etiketi, axis=1)

    # --- 4. Soru grubu sütun ortalamaları ---
    for kat, onekler in GRUP_ONEKLER.items():
        ilgili = [c for c in soru_cols if any(c.startswith(o) for o in onekler)]
        df[f"__{kat}"] = df[ilgili].mean(axis=1) if ilgili else float("nan")

    # --- 5. Ders bazında özet gruplama ---
    gruplama = df.groupby(
        ["Ders Birim", "Ders Kodu", "Kategori"], dropna=False
    ).agg(
        **{"Katılımcı (N)"       : ("Ders Kodu", "count")},
        **{"Ders İçeriği"        : ("__Ders İçeriği",        "mean")},
        **{"Öğretim Elemanı"     : ("__Öğretim Elemanı",     "mean")},
        **{"Ölçme Değerlendirme" : ("__Ölçme Değerlendirme", "mean")},
        **{"Yöntem"              : ("__Yöntem",              "mean")},
        **{"Genel Memnuniyet"    : ("__Genel Memnuniyet",    "mean")},
    ).reset_index()

    # --- 6. Şablonu aç ---
    wb = load_workbook(MD_AKADEMIK_DOSYA)

    # Sadece özet sayfalarını sil, ders detay sayfaları korunacak
    for s in ["Tum_Dersler_Ozet", "Bolum_Genel_Analiz"]:
        if s in wb.sheetnames:
            del wb[s]

    # --- 7. Tüm dersler özet ---
    ws_tum = wb.create_sheet("Tum_Dersler_Ozet")
    tum_dersler_yaz(ws_tum, gruplama)
    print("  ✔ Tum_Dersler_Ozet")

    # --- 8. Bölüm genel analiz ---
    ws_genel = wb.create_sheet("Bolum_Genel_Analiz")
    bolum_genel_yaz(ws_genel, gruplama)
    print("  ✔ Bolum_Genel_Analiz")

    # --- 9. Her ders için detay sayfası ---
    kombinasyonlar = (df.groupby(["Ders Kodu", "Ders Birim"], dropna=False)
                        .size().reset_index().drop(columns=[0]))
    yeni = guncellenen = 0

    for _, r in kombinasyonlar.iterrows():
        ders_k = str(r["Ders Kodu"]).strip()
        birim  = str(r["Ders Birim"]).strip()
        mask   = ((df["Ders Kodu"].astype(str).str.strip() == ders_k) &
                  (df["Ders Birim"].astype(str).str.strip() == birim))
        df_d   = df[mask].copy()
        ad     = sayfa_adi_uret(ders_k, birim)

        if ad in wb.sheetnames:
            ws_d = wb[ad]
            temizle_sayfa(ws_d)
            guncellenen += 1
        else:
            ws_d = wb.create_sheet(ad)
            yeni += 1

        ders_detay_yaz(ws_d, df_d, ders_k, birim, soru_cols)

    print(f"  ✔ Ders detay sayfaları: {guncellenen} güncellendi, {yeni} yeni")

    # --- 10. Kaydet ---
    wb.save(CIKTI_DOSYA)
    print(f"\n[GÖREV 3 TAMAMLANDI] → '{CIKTI_DOSYA}'")
    print(f"  Toplam ders   : {len(kombinasyonlar)}")
    print(f"  Toplam sayfa  : {len(wb.sheetnames)}")


if __name__ == "__main__":
    print("=" * 60)
    print("  GÖREV 3 — MD_Akademik Özet + Ders Detay Sayfaları")
    print("=" * 60)
    md_akademik_ozet_olustur()
    print("\n✔ Görev 3 tamamlandı.")
