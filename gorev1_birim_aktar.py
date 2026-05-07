"""
GÖREV 1 — Başarı Oranı Dosyasından Ders Birim Aktarımı
=======================================================
Başarı oranı dosyasındaki 'Birim' bilgisini,
Ders Kodu + Grup No eşleşmesiyle TBMYO anket dosyasının
'Ders Birim' sütununa yazar.

Çalıştırmadan önce aşağıdaki DOSYA YOLLARI bölümünü güncelleyin.
"""

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# DOSYA YOLLARI  →  kendi bilgisayarınıza göre güncelleyin
# ============================================================
BASARI_DOSYA = "2025-2026 Güz_basari_oranı.xlsx"   # Başarı oranı raporu
TBMYO_DOSYA  = "tbmyo_2025-2026_guz.xlsx"           # Anket ham verisi
CIKTI_DOSYA  = "tbmyo_2025-2026_guz_GUNCELLENDI.xlsx"


def birim_aktar():
    # ----- 1. Dosyaları oku -----
    df_basari = pd.read_excel(BASARI_DOSYA)
    df_tbmyo  = pd.read_excel(TBMYO_DOSYA)

    print(f"Başarı dosyası  : {len(df_basari)} satır")
    print(f"TBMYO dosyası   : {len(df_tbmyo)} satır")

    # ----- 2. Eşleştirme anahtarı -----
    # Grup No float gelebilir (1.0 → 1), int'e çevirerek temizle
    df_basari["_anahtar"] = (
        df_basari["Ders Kodu"].astype(str).str.strip().str.upper()
        + "_"
        + df_basari["Grup No"].astype(str)
          .str.replace(r"\.0$", "", regex=True).str.strip()
    )
    df_tbmyo["_anahtar"] = (
        df_tbmyo["Ders Kodu"].astype(str).str.strip().str.upper()
        + "_"
        + df_tbmyo["Grup No"].astype(str)
          .str.replace(r"\.0$", "", regex=True).str.strip()
    )

    # ----- 3. Birim sözlüğü oluştur -----
    birim_map = df_basari.set_index("_anahtar")["Birim"].to_dict()

    # ----- 4. TBMYO dosyasına yaz (eşleşmeyenler mevcut değerini korur) -----
    before = df_tbmyo["Ders Birim"].notna().sum()
    df_tbmyo["Ders Birim"] = (
        df_tbmyo["_anahtar"].map(birim_map)
        .fillna(df_tbmyo["Ders Birim"])
    )
    df_tbmyo.drop(columns=["_anahtar"], inplace=True)
    after = df_tbmyo["Ders Birim"].notna().sum()

    # ----- 5. Orijinal Excel formatını koruyarak kaydet -----
    wb = load_workbook(TBMYO_DOSYA)
    ws = wb.active

    baslik = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "Ders Birim" not in baslik:
        print("HATA: TBMYO dosyasında 'Ders Birim' sütunu bulunamadı!")
        return None

    birim_col = baslik.index("Ders Birim") + 1  # 1 tabanlı

    for i, deger in enumerate(df_tbmyo["Ders Birim"], start=2):
        ws.cell(row=i, column=birim_col).value = deger

    wb.save(CIKTI_DOSYA)

    # ----- 6. Rapor -----
    eslesen   = df_tbmyo["Ders Birim"].notna().sum()
    eslesmyen = len(df_tbmyo) - eslesen
    print(f"\n[GÖREV 1 TAMAMLANDI] → '{CIKTI_DOSYA}'")
    print(f"  Toplam satır       : {len(df_tbmyo)}")
    print(f"  Önceden dolu       : {before}")
    print(f"  Eşleşip doldurulan : {after - before}")
    print(f"  Hâlâ boş (eşleşme yok): {eslesmyen}")

    return df_tbmyo  # Görev 2'de doğrudan kullanılabilir


if __name__ == "__main__":
    print("=" * 55)
    print("  GÖREV 1 — Ders Birim Aktarımı")
    print("=" * 55)
    birim_aktar()
    print("\n✔ Görev 1 tamamlandı.")
