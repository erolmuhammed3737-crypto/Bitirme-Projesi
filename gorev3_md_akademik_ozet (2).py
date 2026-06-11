"""
GÖREV 3 — MD_Akademik Özet + Ders Detay Sayfaları (Tasarım Bölümü)
====================================================================
DÜZELTİLEN SORUNLAR:
  [9]  Soru gruplama     → Hocamın MD_detay.py'sindeki QUESTION_GROUPS ile birebir
  [10] Özel score mapping → 6_1 "Yapılmadı" → 0,  8_1 "Kaynak önerilmedi" → 0
  [11] Bölüm kapsamı    → HEDEF_BOLUM = "" (boş) → MD.xlsx'teki tüm bölümler işlenir
  [12] Sondaki boşluk   → Ders Birim sütunu str.strip() ile temizleniyor
  [13] Bağımlılık kaldır → MD.xlsx doğrudan okunuyor, Görev 1 çıktısına gerek yok
  [14] Genel Memnuniyet  → 14_1/15_1/16_1 yerine tüm soruların ortalaması

Hocamın QUESTION_GROUPS (MD_detay.py):
    Ders İçeriği          : 1_1, 3_1, 14_1
    Öğretim Elemanı       : 2_1, 4_1, 5_1, 7_1, 9_1, 10_1, 11_1, 12_1, 14_1, 15_1, 16_1
    Ölçme Değerlendirme   : 6_1, 12_1, 13_1, 14_1
    Yöntem                : 4_1, 6_1, 8_1, 10_1, 14_1
    Genel Memnuniyet      : tüm soruların genel ortalaması (calculate_generic_avg)

Çıktı: MD_Akademik_OZET.xlsx  (orijinal MD_Akademik.xlsx bozulmaz)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")


# ============================================================
# DOSYA YOLLARI  →  kendi bilgisayarınıza göre güncelleyin
# ============================================================
# [Düzeltme #13] MD.xlsx doğrudan kaynak — Görev 1 çıktısına bağımlılık kaldırıldı
MD_DOSYA       = "MD.xlsx"          # Tasarım bölümü ham veri
MD_AKADEMIK_DOSYA = "MD_Akademik.xlsx"  # Şablon (mevcut sayfa düzeni korunur)
CIKTI_DOSYA    = "MD_Akademik_OZET.xlsx"

# [Düzeltme #11] Boş bırakırsanız MD.xlsx'teki tüm bölümler işlenir
# Sadece belirli bir bölüm istiyorsanız: HEDEF_BOLUM = "Moda Tasarımı"
HEDEF_BOLUM = ""


# ============================================================
# [Düzeltme #10] LİKERT + ÖZEL CEVAP DÖNÜŞÜMÜ
# Hocamın score_mapping'i (MD_detay.py) ile aynı
# ============================================================
LIKERT_METIN = {
    "kesinlikle katılmıyorum" : 1,
    "katılmıyorum"            : 2,
    "pek fazla katılmıyorum"  : 3,
    "katılıyorum"             : 4,
    "biraz katılıyorum"       : 5,
    "tamamen katılıyorum"     : 6,
}

# Özel 0-puanlı cevaplar (hocamın q6/q8 score_mapping'leri)
OZEL_SIFIR_6_1 = "ödev, proje, ekip çalışması, öğrenci sunumları yapılmadı."
OZEL_SIFIR_8_1 = "ders için kaynak önerilmedi."


def likert(deger, soru_kodu=""):
    """Metin veya sayı Likert değerini 1-6 (ya da 0) sayısına çevirir."""
    if pd.isna(deger):
        return None
    if isinstance(deger, (int, float)):
        v = int(deger)
        return v if 1 <= v <= 6 else None
    if isinstance(deger, str):
        temiz = deger.strip().lower()
        if soru_kodu.startswith("6_1") and temiz == OZEL_SIFIR_6_1:
            return 0
        if soru_kodu.startswith("8_1") and temiz == OZEL_SIFIR_8_1:
            return 0
        return LIKERT_METIN.get(temiz, None)
    return None


# ============================================================
# [Düzeltme #9] SORU GRUPLARI — Hocamın MD_detay.py ile birebir
# NOT: Aynı soru birden fazla kategoride görünebilir (hocamın tasarımı)
# ============================================================
GRUP_ONEKLER = {
    "Ders İçeriği"          : ("1_1", "3_1", "14_1"),
    "Öğretim Elemanı"       : ("2_1", "4_1", "5_1", "7_1", "9_1", "10_1",
                                "11_1", "12_1", "14_1", "15_1", "16_1"),
    "Ölçme Değerlendirme"   : ("6_1", "12_1", "13_1", "14_1"),
    "Yöntem"                : ("4_1", "6_1", "8_1", "10_1", "14_1"),
}
KAT_LISTESI = list(GRUP_ONEKLER.keys())


# ============================================================
# BİÇİMLENDİRME SABİTLERİ
# ============================================================
R_BASLIK = PatternFill("solid", fgColor="1F4E79")
R_HDR    = PatternFill("solid", fgColor="2E75B6")
R_ALT    = PatternFill("solid", fgColor="D6E4F0")
R_BEYAZ  = PatternFill("solid", fgColor="FFFFFF")
R_TOPLAM = PatternFill("solid", fgColor="FCE4D6")
ince     = Side(style="thin", color="B0B0B0")
KENAR    = Border(left=ince, right=ince, top=ince, bottom=ince)
ORTALA   = Alignment(horizontal="center", vertical="center", wrap_text=True)
SOLA     = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def yaz(cell, val, fill=R_BEYAZ, bold=False, fmt=None, align=ORTALA, fsize=10):
    cell.value     = val
    cell.fill      = fill
    cell.font      = Font(name="Arial", bold=bold, size=fsize,
                          color="FFFFFF" if fill in (R_BASLIK, R_HDR) else "000000")
    cell.alignment = align
    cell.border    = KENAR
    if fmt:
        cell.number_format = fmt


# ============================================================
# YARDIMCI FONKSİYONLAR
# ============================================================
def kategori_etiketi(satir):
    birim   = str(satir.get("Ders Birim", "") or "").strip()  # [#12] strip()
    uzaktan = "uzaktan" in birim.lower()
    kod     = str(satir.get("Ders Kodu", "") or "")
    rakam   = next((c for c in kod if c.isdigit()), "1")
    try:    sinif = min(int(rakam), 2)
    except: sinif = 1
    return f"{sinif}. Sınıf - {'Uzaktan Öğretim' if uzaktan else 'Birinci Öğretim'}"


def sayfa_adi_uret(ders_kodu, birim):
    """DersKodu_TIP_BolumAdi  (TIP: OR=Örgün, UZ=Uzaktan)"""
    tip        = "UZ" if "uzaktan" in str(birim).lower() else "OR"
    bolum_temiz = (str(birim).strip()            # [#12] strip()
                   .replace("(Uzaktan Öğretim)", "")
                   .replace("(İÖ)", "")
                   .strip())
    raw = f"{ders_kodu}_{tip}_{bolum_temiz}"
    for ch in ("/", "\\", "?", "*", "[", "]", ":"):
        raw = raw.replace(ch, "-")
    return raw[:31]


def temizle_sayfa(ws):
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))
    for row in ws.iter_rows():
        for cell in row:
            try:
                cell.value         = None
                cell.fill          = PatternFill()
                cell.font          = Font()
                cell.border        = Border()
                cell.number_format = "General"
            except AttributeError:
                pass


# ============================================================
# [Düzeltme #10] KATEGORI ORTALAMASI HESABI
# Hocamın calculate_group_averages mantığı:
# Aynı soru birden fazla kategoride olabilir — her soru sayılır
# ============================================================
def kategori_ort(df_ders, soru_cols, onekler):
    """Belirtilen önek listesiyle eşleşen sütunların global ortalaması."""
    ilgili = [c for c in soru_cols if any(c.startswith(o) for o in onekler)]
    if not ilgili:
        return float("nan")
    toplam = 0.0
    sayi   = 0
    for col in ilgili:
        vals = df_ders[col].dropna()
        toplam += vals.sum()
        sayi   += len(vals)
    return toplam / sayi if sayi > 0 else float("nan")


# ============================================================
# [Düzeltme #14] GENEL MEMNUNİYET  —  tüm soruların ortalaması
# Hocamın calculate_generic_avg mantığı
# ============================================================
def genel_ort(df_ders, soru_cols):
    toplam = 0.0
    sayi   = 0
    for col in soru_cols:
        vals = df_ders[col].dropna()
        toplam += vals.sum()
        sayi   += len(vals)
    return round(toplam / sayi, 2) if sayi > 0 else float("nan")


# ============================================================
# 1. DERS DETAY SAYFASI
# ============================================================
def ders_detay_yaz(ws, df_ders, ders_kodu, birim, soru_cols):
    ws.sheet_view.showGridLines = False
    n = len(df_ders)

    ws.merge_cells("A1:B1")
    yaz(ws["A1"], f"{ders_kodu}  |  {birim}",
        fill=R_BASLIK, bold=True, fsize=11, align=SOLA)
    ws.merge_cells("D1:E1")
    yaz(ws["D1"], "Soru Bazlı Ortalamalar",
        fill=R_BASLIK, bold=True, fsize=11)
    ws.row_dimensions[1].height = 28

    yaz(ws["A2"], "Kategori",  fill=R_HDR, bold=True)
    yaz(ws["B2"], "Ortalama",  fill=R_HDR, bold=True)
    yaz(ws["D2"], "Soru",      fill=R_HDR, bold=True, align=SOLA)
    yaz(ws["E2"], "Ortalama",  fill=R_HDR, bold=True)
    ws.row_dimensions[2].height = 22

    # Sol: Kategori ortalamaları — [#9] hocamın grupları
    r = 3
    for kat, onekler in GRUP_ONEKLER.items():
        ort   = kategori_ort(df_ders, soru_cols, onekler)
        dolgu = R_ALT if r % 2 == 0 else R_BEYAZ
        yaz(ws.cell(r, 1), kat, dolgu, align=SOLA)
        yaz(ws.cell(r, 2), round(ort, 2) if pd.notna(ort) else "-", dolgu, fmt="0.00")
        r += 1

    # [Düzeltme #14] Genel Memnuniyet = tüm soruların ortalaması
    gen = genel_ort(df_ders, soru_cols)
    yaz(ws.cell(r, 1), "Genel Memnuniyet Ortalaması", R_TOPLAM, bold=True, align=SOLA)
    yaz(ws.cell(r, 2), round(gen, 2) if pd.notna(gen) else "-",
        R_TOPLAM, bold=True, fmt="0.00")
    r += 2

    yaz(ws.cell(r, 1), f"Katılımcı Sayısı (N): {n}", R_BEYAZ, bold=True, align=SOLA)
    ws.cell(r, 2).fill = R_BEYAZ

    # Sağ: Soru bazlı ortalamalar
    for i, soru_col in enumerate(soru_cols):
        row_idx = i + 3
        dolgu   = R_ALT if row_idx % 2 == 0 else R_BEYAZ
        ort_s   = df_ders[soru_col].mean()
        yaz(ws.cell(row_idx, 4), soru_col.strip(), dolgu, align=SOLA)
        yaz(ws.cell(row_idx, 5),
            round(ort_s, 2) if pd.notna(ort_s) else "-", dolgu, fmt="0.00")
        ws.row_dimensions[row_idx].height = 16

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 78
    ws.column_dimensions["E"].width = 11
    ws.freeze_panes = "A3"


# ============================================================
# 2. TÜM DERSLER ÖZET
# ============================================================
def tum_dersler_yaz(ws, gruplama):
    ws.sheet_view.showGridLines = False
    SUTUNLAR = [
        "Ders Birim", "Ders Kodu", "Kategori", "Katılımcı (N)",
        "Ders İçeriği", "Öğretim Elemanı", "Ölçme Değerlendirme",
        "Yöntem", "Genel Memnuniyet"
    ]

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value, c.fill = "Tüm Dersler – Ders Değerlendirme Özeti", R_BASLIK
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.alignment = ORTALA
    c.border    = KENAR
    ws.row_dimensions[1].height = 32

    for ci, ad in enumerate(SUTUNLAR, 1):
        h = ws.cell(2, ci)
        h.value, h.fill = ad, R_HDR
        h.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        h.alignment = ORTALA
        h.border    = KENAR
    ws.row_dimensions[2].height = 36

    for ri, row in enumerate(gruplama.itertuples(index=False), start=3):
        dolgu = R_ALT if ri % 2 == 0 else R_BEYAZ
        for ci, val in enumerate(row, 1):
            if ci in (5, 6, 7, 8, 9):
                yaz(ws.cell(ri, ci),
                    round(float(val), 2) if pd.notna(val) else "", dolgu, fmt="0.00")
            elif ci in (1, 3):
                yaz(ws.cell(ri, ci), val, dolgu, align=SOLA)
            else:
                yaz(ws.cell(ri, ci), val, dolgu)
        ws.row_dimensions[ri].height = 18

    son = 2 + len(gruplama)
    if len(gruplama) > 0:
        r = son + 1
        ws.row_dimensions[r].height = 20
        yaz(ws.cell(r, 1), "GENEL ORTALAMA", R_TOPLAM, bold=True, align=SOLA)
        yaz(ws.cell(r, 2), "", R_TOPLAM)
        yaz(ws.cell(r, 3), "", R_TOPLAM)
        yaz(ws.cell(r, 4), int(gruplama["Katılımcı (N)"].sum()), R_TOPLAM, bold=True)
        for ci, sc in enumerate(KAT_LISTESI + ["Genel Memnuniyet"], 5):
            yaz(ws.cell(r, ci), round(gruplama[sc].mean(), 2),
                R_TOPLAM, bold=True, fmt="0.00")

    for ci, g in enumerate([30, 12, 28, 14, 13, 16, 20, 12, 18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = g
    ws.freeze_panes = "A3"


# ============================================================
# 3. BÖLÜM GENEL ANALİZ
# ============================================================
def bolum_genel_yaz(ws, gruplama):
    ws.sheet_view.showGridLines = False
    SUTUNLAR = ["Kategori"] + KAT_LISTESI + ["Genel Memnuniyet"]

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value, c.fill = "Bölüm Genel Analizi – Kategori Ortalamaları", R_BASLIK
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.alignment = ORTALA
    c.border    = KENAR
    ws.row_dimensions[1].height = 28

    for ci, ad in enumerate(SUTUNLAR, 1):
        h = ws.cell(2, ci)
        h.value, h.fill = ad, R_HDR
        h.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        h.alignment = ORTALA
        h.border    = KENAR
    ws.row_dimensions[2].height = 36

    kat_ozet  = gruplama.groupby("Kategori")[KAT_LISTESI + ["Genel Memnuniyet"]].mean()
    kat_sirali = sorted(kat_ozet.index.tolist())
    kat_ozet   = kat_ozet.loc[kat_sirali]

    for ri, (kat, row) in enumerate(kat_ozet.iterrows(), start=3):
        dolgu = R_ALT if ri % 2 == 0 else R_BEYAZ
        yaz(ws.cell(ri, 1), kat, dolgu, align=SOLA)
        for ci, sc in enumerate(KAT_LISTESI + ["Genel Memnuniyet"], 2):
            yaz(ws.cell(ri, ci),
                round(float(row[sc]), 2) if pd.notna(row[sc]) else "",
                dolgu, fmt="0.00")
        ws.row_dimensions[ri].height = 20

    r = 2 + len(kat_ozet) + 1
    ws.row_dimensions[r].height = 20
    yaz(ws.cell(r, 1), "GENEL ORTALAMA", R_TOPLAM, bold=True, align=SOLA)
    for ci, sc in enumerate(KAT_LISTESI + ["Genel Memnuniyet"], 2):
        yaz(ws.cell(r, ci), round(gruplama[sc].mean(), 2),
            R_TOPLAM, bold=True, fmt="0.00")

    ws.column_dimensions["A"].width = 32
    for ci in range(2, 8):
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.freeze_panes = "A3"


# ============================================================
# ANA FONKSİYON
# ============================================================
def md_akademik_ozet_olustur():
    # --- 1. [Düzeltme #13] MD.xlsx'ten doğrudan oku ---
    df = pd.read_excel(MD_DOSYA)

    # [Düzeltme #12] Ders Birim sondaki/baştaki boşlukları temizle
    df.columns = (df.columns
                  .str.strip()
                  .str.replace("\u200b", "", regex=False)
                  .str.replace("\xa0", "",  regex=False))
    if "Ders Birim" in df.columns:
        df["Ders Birim"] = df["Ders Birim"].astype(str).str.strip()

    print(f"Kaynak: {MD_DOSYA}  ({len(df)} satır)")

    # --- 2. [Düzeltme #11] Bölüm filtresi — boş → hepsi ---
    if HEDEF_BOLUM:
        mask = df["Ders Birim"].str.contains(HEDEF_BOLUM.strip(), na=False)
        df   = df[mask].copy()
        print(f"Filtre ('{HEDEF_BOLUM}'): {len(df)} satır")
    else:
        print(f"Filtre yok — tüm bölümler işleniyor.")

    # --- 3. [Düzeltme #10] Likert + özel mapping dönüşümü ---
    soru_cols = [c for c in df.columns if str(c)[0].isdigit()]
    for s in soru_cols:
        soru_kodu = s.split()[0] if " " in s else s
        df[s] = df[s].apply(lambda v, sk=soru_kodu: likert(v, sk))

    # --- 4. Kategori etiketi ---
    df["Kategori"] = df.apply(kategori_etiketi, axis=1)

    # --- 5. [Düzeltme #9] Kategori sütunları (hocamın QUESTION_GROUPS) ---
    for kat, onekler in GRUP_ONEKLER.items():
        ilgili = [c for c in soru_cols if any(c.startswith(o) for o in onekler)]
        if ilgili:
            # Hocamın hesabı: tüm ilgili soruların global ortalaması
            def kat_ort_satir(row, cols=ilgili):
                vals = row[cols].dropna()
                return vals.mean() if len(vals) > 0 else float("nan")
            df[f"__{kat}"] = df.apply(kat_ort_satir, axis=1)
        else:
            df[f"__{kat}"] = float("nan")

    # --- 6. [Düzeltme #14] Genel Memnuniyet = tüm soruların ortalaması ---
    df["__Genel Memnuniyet"] = df[soru_cols].mean(axis=1)

    # --- 7. Ders bazında özet gruplama ---
    gruplama = df.groupby(
        ["Ders Birim", "Ders Kodu", "Kategori"], dropna=False
    ).agg(
        **{"Katılımcı (N)"       : ("Ders Kodu",               "count")},
        **{"Ders İçeriği"        : ("__Ders İçeriği",          "mean")},
        **{"Öğretim Elemanı"     : ("__Öğretim Elemanı",       "mean")},
        **{"Ölçme Değerlendirme" : ("__Ölçme Değerlendirme",   "mean")},
        **{"Yöntem"              : ("__Yöntem",                 "mean")},
        **{"Genel Memnuniyet"    : ("__Genel Memnuniyet",       "mean")},
    ).reset_index()

    # --- 8. Şablonu aç ---
    wb = load_workbook(MD_AKADEMIK_DOSYA)

    # Sadece özet sayfaları sil, ders detay sayfaları korunur
    for s in ["Tum_Dersler_Ozet", "Bolum_Genel_Analiz"]:
        if s in wb.sheetnames:
            del wb[s]

    # --- 9. Özet sayfaları oluştur ---
    ws_tum = wb.create_sheet("Tum_Dersler_Ozet")
    tum_dersler_yaz(ws_tum, gruplama)
    print("  ✔ Tum_Dersler_Ozet")

    ws_genel = wb.create_sheet("Bolum_Genel_Analiz")
    bolum_genel_yaz(ws_genel, gruplama)
    print("  ✔ Bolum_Genel_Analiz")

    # --- 10. [Düzeltme #11] Her bölüm için ders detay sayfaları ---
    kombinasyonlar = (df.groupby(["Ders Kodu", "Ders Birim"], dropna=False)
                        .size().reset_index().drop(columns=[0]))
    yeni = guncellenen = 0

    for _, r in kombinasyonlar.iterrows():
        ders_k = str(r["Ders Kodu"]).strip()
        birim  = str(r["Ders Birim"]).strip()    # [#12] strip()
        mask   = (
            (df["Ders Kodu"].astype(str).str.strip() == ders_k)
            & (df["Ders Birim"].astype(str).str.strip() == birim)
        )
        df_d = df[mask].copy()
        ad   = sayfa_adi_uret(ders_k, birim)

        if ad in wb.sheetnames:
            ws_d = wb[ad]
            temizle_sayfa(ws_d)
            guncellenen += 1
        else:
            ws_d = wb.create_sheet(ad)
            yeni += 1

        ders_detay_yaz(ws_d, df_d, ders_k, birim, soru_cols)

    print(f"  ✔ Ders detay sayfaları: {guncellenen} güncellendi, {yeni} yeni")

    # --- 11. Kaydet ---
    wb.save(CIKTI_DOSYA)
    print(f"\n[GÖREV 3 TAMAMLANDI] → '{CIKTI_DOSYA}'")
    print(f"  Toplam ders   : {len(kombinasyonlar)}")
    print(f"  Toplam sayfa  : {len(wb.sheetnames)}")


if __name__ == "__main__":
    print("=" * 60)
    print("  GÖREV 3 — MD_Akademik Özet + Ders Detay Sayfaları")
    print("=" * 60)
    md_akademik_ozet_olustur()
    print("\n✔ Görev 3 tamamlandı.")
