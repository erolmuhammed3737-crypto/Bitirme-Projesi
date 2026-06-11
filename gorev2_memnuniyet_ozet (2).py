"""
GÖREV 2 — Memnuniyet Ortalamaları & Özet Sayfaları
====================================================
DÜZELTİLEN SORUNLAR:
  [4] Giriş formatı  → Tek workbook yerine ANKET_KLASOR'daki her .xlsx ayrı dönem
  [5] İki anket formu → 14_1 içeriğine bakarak eski/yeni form otomatik algılanır
       ESKİ (≤23/24 Güz): 14_1 = devam sorusu, 15_1 = tekrar sayısı → Likert değil
       YENİ (≥23/24 Bahar): 14_1 = "beklentilerimi karşıladı" → geçerli Likert
  [6] Soru gruplama  → Hocamın MD_detay.py referansıyla eşleştirildi (yeni form)
  [7] Özel cevaplar  → 6_1 "Yapılmadı" → 0 ve 8_1 "Kaynak önerilmedi" → 0
  [8] PNG silme sırası → os.unlink() wb.save() dan SONRA çağrılıyor

Kullanım:
  • ANKET_KLASOR'u memnuniyet xlsx dosyalarının bulunduğu klasöre ayarlayın
  • python gorev2_memnuniyet_ozet.py
"""

import os
import glob
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import tempfile
import warnings
warnings.filterwarnings("ignore")


# ============================================================
# DOSYA YOLLARI  →  kendi bilgisayarınıza göre güncelleyin
# ============================================================
ANKET_KLASOR  = "."    # Tüm dönem xlsx'lerinin bulunduğu klasör
                        # (ör. r"C:\...\Memnuniyet Anketleri")
CIKTI_KLASOR  = "."    # Çıktı klasörü

# Hangi dosya adı kalıbı dönem xlsx'i?
DOSYA_KALIP   = "*.xlsx"


# ============================================================
# LİKERT VE ÖZEL CEVAP DÖNÜŞÜM HARİTALARI
# ============================================================
# Temel 1-6 Likert (yeni form metin değerler için)
LIKERT_METIN = {
    "kesinlikle katılmıyorum" : 1,
    "katılmıyorum"            : 2,
    "pek fazla katılmıyorum"  : 3,
    "katılıyorum"             : 4,
    "biraz katılıyorum"       : 5,
    "tamamen katılıyorum"     : 6,
}

# [Düzeltme #7] Özel 0-puanlı cevaplar
# 6_1: ders harici yöntemler uygulanmadıysa → 0
# 8_1: kaynak önerilmediyse → 0
OZEL_SIFIR_6_1 = "ödev, proje, ekip çalışması, öğrenci sunumları yapılmadı."
OZEL_SIFIR_8_1 = "ders için kaynak önerilmedi."


def likert_donustur(deger, soru_kodu=""):
    """Metin veya sayı Likert değerini 1-6 arası sayıya çevirir."""
    if pd.isna(deger):
        return None
    if isinstance(deger, (int, float)):
        v = int(deger)
        return v if 1 <= v <= 6 else None
    if isinstance(deger, str):
        temiz = deger.strip().lower()
        # Özel 0-mapping kontrolü
        if soru_kodu.startswith("6_1") and temiz == OZEL_SIFIR_6_1:
            return 0
        if soru_kodu.startswith("8_1") and temiz == OZEL_SIFIR_8_1:
            return 0
        return LIKERT_METIN.get(temiz, None)
    return None


# ============================================================
# [Düzeltme #5] FORM TESPİTİ
# 14_1 sütun adına bakarak eski/yeni form belirlenir.
# ============================================================
YENI_FORM_ANAHTAR = "beklentilerimi"   # Yeni formda 14_1 bu kelimeyi içerir


def form_tespit_et(df):
    """
    df sütunlarına bakarak 'eski' veya 'yeni' döndürür.
    YENİ: 14_1 = "Ders genel olarak beklentilerimi karşıladı."
    ESKİ: 14_1 = "Bu derse ne derece devam ettiniz?"
    """
    q14_sutun = next((c for c in df.columns if str(c).startswith("14_1")), None)
    if q14_sutun and YENI_FORM_ANAHTAR in str(q14_sutun).lower():
        return "yeni"
    return "eski"


# ============================================================
# [Düzeltme #6] SORU GRUPLARI  —  form tipine göre
# YENİ FORM: Hocamın MD_detay.py referansı
# ESKİ FORM: 14_1/15_1 (devam/tekrar) hariç uyarlanmış gruplama
# ============================================================

# Yeni form (≥23/24 Bahar) — hocamın MD_detay.py'siyle birebir
GRUP_ONEKLER_YENI = {
    "Ders İçeriği"          : ("1_1", "3_1", "14_1"),
    "Öğretim Elemanı"       : ("2_1", "4_1", "5_1", "7_1", "9_1", "10_1", "11_1",
                                "12_1", "14_1", "15_1", "16_1"),
    "Ölçme Değerlendirme"   : ("6_1", "12_1", "13_1", "14_1"),
    "Yöntem"                : ("4_1", "6_1", "8_1", "10_1", "14_1"),
}
# Yeni formda Genel Memnuniyet = tüm 16 sorunun ortalaması (hocamın calculate_generic_avg)
YENI_FORM_TOPLAM_SORU_SAYISI = 16

# Eski form (≤23/24 Güz) — 14_1 ve 15_1 Likert değil, dışarıda tutulur
# (Bu dönemler için 1_1→13_1 arası 13 soru gerçek Likert)
GRUP_ONEKLER_ESKI = {
    "Ders İçeriği"          : ("1_1",),
    "Öğretim Elemanı"       : ("2_1", "3_1", "5_1", "6_1", "7_1", "12_1"),
    "Ölçme Değerlendirme"   : ("10_1", "11_1"),
    "Yöntem"                : ("4_1", "8_1", "9_1"),
}
# Eski formda Genel Memnuniyet = 1_1 ile 13_1 arasındaki sorular (devam/tekrar hariç)
ESKI_FORM_HARIC_ONEKLER = ("14_1", "15_1", "16_1")  # Likert olmayan sütunlar


KATEGORI_SUTUNLAR = ["Ders İçeriği", "Öğretim Elemanı", "Ölçme Değerlendirme", "Yöntem"]


# ============================================================
# EXCEL BİÇİMLENDİRME SABİTLERİ
# ============================================================
RENK_BASLIK    = PatternFill("solid", fgColor="1F4E79")
RENK_SUTUN_HDR = PatternFill("solid", fgColor="2E75B6")
RENK_SATIR_ALT = PatternFill("solid", fgColor="D6E4F0")
RENK_BEYAZ     = PatternFill("solid", fgColor="FFFFFF")
RENK_TOPLAM    = PatternFill("solid", fgColor="FCE4D6")
ince           = Side(style="thin", color="B0B0B0")
KENAR          = Border(left=ince, right=ince, top=ince, bottom=ince)
ORTALA         = Alignment(horizontal="center", vertical="center", wrap_text=True)
SOLA           = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def h_yaz(cell, val, fill=RENK_BEYAZ, bold=False, fmt=None, align=ORTALA):
    cell.value     = val
    cell.fill      = fill
    cell.font      = Font(name="Arial", bold=bold, size=10,
                          color="FFFFFF" if fill == RENK_BASLIK else "000000")
    cell.alignment = align
    cell.border    = KENAR
    if fmt:
        cell.number_format = fmt


# ============================================================
# KATEGORİ HESAPLAMA
# ============================================================
def kategori_belirle(satir):
    birim   = str(satir.get("Ders Birim", "") or "")
    uzaktan = "uzaktan" in birim.lower()
    kod     = str(satir.get("Ders Kodu", "") or "")
    rakam   = next((c for c in kod if c.isdigit()), "1")
    try:
        sinif = min(int(rakam), 2)
    except ValueError:
        sinif = 1
    ogretim = "Uzaktan Öğretim" if uzaktan else "Birinci Öğretim"
    return f"{sinif}. Sınıf - {ogretim}"


def hesapla_gruplama(df, form_tipi):
    """
    Ham anket df'inden ders bazında kategori ortalamaları hesaplar.
    form_tipi: 'yeni' veya 'eski'
    """
    grup_onekler = GRUP_ONEKLER_YENI if form_tipi == "yeni" else GRUP_ONEKLER_ESKI

    # Soru sütunlarını bul
    tum_soru_cols = [c for c in df.columns if str(c)[0].isdigit()]

    # [Düzeltme #5] Eski formda Likert olmayan sütunları dışla
    if form_tipi == "eski":
        soru_cols = [c for c in tum_soru_cols
                     if not any(c.startswith(p) for p in ESKI_FORM_HARIC_ONEKLER)]
    else:
        soru_cols = tum_soru_cols

    # [Düzeltme #7] Özel 0-mapping dahil Likert dönüşümü
    for s in soru_cols:
        soru_kodu = s.split()[0] if " " in s else s
        df[s] = df[s].apply(lambda v, sk=soru_kodu: likert_donustur(v, sk))

    # Kategori etiketleri
    df["Kategori"] = df.apply(kategori_belirle, axis=1)

    # Kategori grubu sütun ortalamaları
    for kat, onekler in grup_onekler.items():
        ilgili = [c for c in soru_cols if any(c.startswith(o) for o in onekler)]
        df[f"__{kat}"] = df[ilgili].mean(axis=1) if ilgili else float("nan")

    # [Düzeltme #6] Genel Memnuniyet
    # Yeni form: 1_1'den 16_1'e kadar tüm soruların ortalaması
    # Eski form: Likert olan soruların (1_1–13_1) ortalaması
    df["__Genel Memnuniyet"] = df[soru_cols].mean(axis=1)

    # Ders Birim + Ders Kodu + Kategori bazında grupla
    gruplama = df.groupby(
        ["Ders Birim", "Ders Kodu", "Kategori"], dropna=False
    ).agg(
        **{"Katılımcı (N)"       : ("Ders Kodu",               "count")},
        **{"Ders İçeriği"        : ("__Ders İçeriği",          "mean")},
        **{"Öğretim Elemanı"     : ("__Öğretim Elemanı",       "mean")},
        **{"Ölçme Değerlendirme" : ("__Ölçme Değerlendirme",   "mean")},
        **{"Yöntem"              : ("__Yöntem",                 "mean")},
        **{"Genel Memnuniyet"    : ("__Genel Memnuniyet",       "mean")},
    ).reset_index()

    return gruplama


# ============================================================
# ÖZET SAYFA YAZICI
# ============================================================
def ozet_sayfa_yaz(ws, df_bolum, baslik_str):
    SUTUNLAR = [
        "Ders Birim", "Ders Kodu", "Kategori", "Katılımcı (N)",
        "Ders İçeriği", "Öğretim Elemanı", "Ölçme Değerlendirme",
        "Yöntem", "Genel Memnuniyet"
    ]

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value, c.fill = f"{baslik_str} – Ders Değerlendirme Özeti", RENK_BASLIK
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.alignment = ORTALA
    c.border    = KENAR
    ws.row_dimensions[1].height = 32

    for ci, ad in enumerate(SUTUNLAR, 1):
        h = ws.cell(2, ci)
        h.value, h.fill = ad, RENK_SUTUN_HDR
        h.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        h.alignment = ORTALA
        h.border    = KENAR
    ws.row_dimensions[2].height = 36

    for ri, row in enumerate(df_bolum.itertuples(index=False), start=3):
        dolgu = RENK_SATIR_ALT if ri % 2 == 0 else RENK_BEYAZ
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            if ci in (5, 6, 7, 8, 9):
                h_yaz(cell, round(float(val), 2) if pd.notna(val) else "", dolgu, fmt="0.00")
            elif ci in (1, 3):
                h_yaz(cell, val, dolgu, align=SOLA)
            else:
                h_yaz(cell, val, dolgu)
        ws.row_dimensions[ri].height = 18

    son = 2 + len(df_bolum)
    if len(df_bolum) > 0:
        r = son + 1
        ws.row_dimensions[r].height = 20
        h_yaz(ws.cell(r, 1), "BÖLÜM ORTALAMASI", RENK_TOPLAM, bold=True, align=SOLA)
        h_yaz(ws.cell(r, 2), "", RENK_TOPLAM)
        h_yaz(ws.cell(r, 3), "", RENK_TOPLAM)
        h_yaz(ws.cell(r, 4), int(df_bolum["Katılımcı (N)"].sum()), RENK_TOPLAM, bold=True)
        for ci, sc in enumerate(KATEGORI_SUTUNLAR + ["Genel Memnuniyet"], 5):
            h_yaz(ws.cell(r, ci),
                  round(df_bolum[sc].mean(), 2), RENK_TOPLAM, bold=True, fmt="0.00")

    for ci, g in enumerate([30, 12, 28, 14, 13, 16, 20, 12, 18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = g
    ws.freeze_panes = "A3"


# ============================================================
# GRAFİK OLUŞTURUCU  —  Sınıf bazında ayrı
# ============================================================
def grafik_olustur(gruplama, sinif_no, donem_adi):
    prefix  = f"{sinif_no}. Sınıf"
    df_plot = gruplama[gruplama["Kategori"].str.startswith(prefix)].copy()

    if df_plot.empty:
        return None

    fig, axes = plt.subplots(
        1, 2,
        figsize=(16, max(5, len(df_plot) * 0.45 + 2)),
        sharey=False
    )
    fig.suptitle(
        f"{donem_adi}  |  {sinif_no}. Sınıf Ders Memnuniyet Ortalamaları",
        fontsize=13, fontweight="bold", y=1.01
    )

    for ax, tur in zip(axes, ["Birinci Öğretim", "Uzaktan Öğretim"]):
        df_tur = df_plot[df_plot["Kategori"].str.contains(tur)].copy()
        ax.set_title(tur, fontsize=11, fontweight="bold", pad=8)
        ax.set_xlim(1, 6)
        ax.axvline(x=3.5, color="gray", linestyle="--", linewidth=0.8, alpha=0.6)

        if df_tur.empty:
            ax.text(0.5, 0.5, "Veri yok", ha="center", va="center",
                    transform=ax.transAxes, fontsize=10, color="gray")
            ax.set_xlabel("Ortalama (1-6)")
            continue

        etketler = (df_tur["Ders Kodu"].astype(str)
                    + "\n" + df_tur["Ders Birim"].astype(str).str[:25])
        x    = range(len(df_tur))
        w    = 0.14

        for i, (kat, renk) in enumerate([
            ("Ders İçeriği",        "#1F4E79"),
            ("Öğretim Elemanı",     "#2E75B6"),
            ("Ölçme Değerlendirme", "#70AD47"),
            ("Yöntem",              "#FFC000"),
            ("Genel Memnuniyet",    "#ED7D31"),
        ]):
            offs = [xi + (i - 2) * w for xi in x]
            vals = df_tur[kat].fillna(0).tolist()
            ax.barh(offs, vals, height=w * 0.9,
                    label=kat, color=renk, alpha=0.88)
            for ox, v in zip(offs, vals):
                if v > 0:
                    ax.text(v + 0.04, ox, f"{v:.2f}", va="center",
                            fontsize=6.5, color="#333333")

        ax.set_yticks(list(x))
        ax.set_yticklabels(etketler, fontsize=7.5)
        ax.set_xlabel("Ortalama (1-6)", fontsize=9)
        ax.legend(loc="lower right", fontsize=7, framealpha=0.7)
        ax.grid(axis="x", linestyle=":", alpha=0.5)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    plt.tight_layout()

    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    plt.savefig(tmp.name, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return tmp.name


def grafik_sayfasi_ekle(wb, png_yolu, sayfa_adi):
    ws = wb.create_sheet(sayfa_adi)
    ws.column_dimensions["A"].width  = 120
    ws.row_dimensions[1].height      = 600
    img         = XLImage(png_yolu)
    img.anchor  = "A1"
    ws.add_image(img)


# ============================================================
# TEK DÖNEM İŞLE
# ============================================================
def isle_donem(dosya_yolu, cikti_klasor="."):
    donem_adi = os.path.splitext(os.path.basename(dosya_yolu))[0]
    df_ham    = pd.read_excel(dosya_yolu)

    # Sütun adlarını temizle (boşluk, görünmez karakter)
    df_ham.columns = (
        df_ham.columns
              .str.strip()
              .str.replace("\u200b", "", regex=False)
              .str.replace("\xa0", "",  regex=False)
    )

    # [Düzeltme #5] Form tipini tespit et
    form_tipi = form_tespit_et(df_ham)
    print(f"\n  ► {donem_adi}  ({len(df_ham)} satır, {form_tipi.upper()} form)")

    gruplama = hesapla_gruplama(df_ham.copy(), form_tipi)

    # Yeni workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Tüm dersler özet
    ozet_sayfa_yaz(wb.create_sheet("Tum_Dersler_Ozet"), gruplama, "Tüm Dersler")

    # Bölüm sayfaları
    bolumler = sorted(gruplama["Ders Birim"].dropna().unique())
    for bolum in bolumler:
        df_b  = gruplama[gruplama["Ders Birim"] == bolum].copy()
        guvli = (str(bolum)
                 .replace("/", "-").replace("\\", "-")
                 .replace("?", "").replace("*", "")
                 .replace("[", "").replace("]", "").replace(":", "-"))
        sayfa = f"Bolum_{guvli}"[:31]
        ozet_sayfa_yaz(wb.create_sheet(sayfa), df_b, bolum)

    # Grafik PNG'leri — [Düzeltme #8] unlink KAYDEDILDIKTEN SONRA
    png_listesi = []
    for sinif in [1, 2]:
        png = grafik_olustur(gruplama, sinif, donem_adi)
        if png:
            grafik_sayfasi_ekle(wb, png, f"Sinif{sinif}_Grafik")
            png_listesi.append(png)

    # Kaydet
    cikti = os.path.join(cikti_klasor, f"{donem_adi}_OZET.xlsx")
    wb.save(cikti)

    # [Düzeltme #8] PNG artık kaydedildikten sonra siliniyor
    for png in png_listesi:
        try:
            os.unlink(png)
        except OSError:
            pass

    print(f"    ✔ Kaydedildi → {cikti}")
    print(f"      Bölüm sayısı : {len(bolumler)}")
    print(f"      Toplam ders  : {len(gruplama)}")
    return cikti


# ============================================================
# ANA FONKSİYON  —  [Düzeltme #4] Klasördeki her xlsx = bir dönem
# ============================================================
def isle_klasor(klasor=None, cikti_klasor=None):
    klasor       = klasor       or ANKET_KLASOR
    cikti_klasor = cikti_klasor or CIKTI_KLASOR

    print("=" * 60)
    print("  GÖREV 2 — Memnuniyet Ortalamaları & Özet Sayfaları")
    print("=" * 60)

    dosyalar = sorted(glob.glob(os.path.join(klasor, DOSYA_KALIP)))
    # Çıktı dosyalarını tekrar okumamak için filtrele
    dosyalar = [d for d in dosyalar if "_OZET" not in os.path.basename(d)]

    if not dosyalar:
        print(f"\nHATA: '{klasor}' klasöründe xlsx dosyası bulunamadı.")
        return

    print(f"\nKlasör : {klasor}")
    print(f"Bulunan: {len(dosyalar)} dosya")

    olusturulan = []
    for d in dosyalar:
        try:
            cikti = isle_donem(d, cikti_klasor)
            olusturulan.append(cikti)
        except Exception as e:
            print(f"\n  HATA ({os.path.basename(d)}): {e}")

    print(f"\n{'=' * 60}")
    print(f"✔ Toplam {len(olusturulan)} dönem işlendi.")
    for f in olusturulan:
        print(f"   • {f}")


if __name__ == "__main__":
    isle_klasor()
