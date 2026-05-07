"""
GÖREV 2 — Memnuniyet Ortalamaları & Özet Sayfaları
====================================================
• GİRİŞ  : Birden fazla dönemi içeren tek bir workbook
            (her sheet = bir dönem, ör. 24_25_G / 24_25_B / 25_26_G)
• ÇIKIŞ  : Her dönem için ayrı bir workbook
            → <donem>_OZET.xlsx   (ör. 25_26_G_OZET.xlsx)

Özellikler
----------
• Metin veya sayı formatındaki Likert yanıtlarını otomatik dönüştürür
• 1. Sınıf ve 2. Sınıf için AYRI grafikler oluşturur
• Birinci Öğretim / Uzaktan Öğretim ayrımını korur
• Her çıktı workbook'a şunlar eklenir:
    - Tum_Dersler_Ozet   → tüm bölümler
    - Bolum_<BolumAdi>   → her bölüm için ayrı sayfa
    - Sinif1_Grafik       → 1. sınıf grafik sayfası
    - Sinif2_Grafik       → 2. sınıf grafik sayfası

Çalıştırmadan önce DOSYA YOLLARI bölümünü güncelleyin.
"""

import os
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import tempfile
import warnings
warnings.filterwarnings("ignore")


# ============================================================
# DOSYA YOLLARI  →  kendi bilgisayarınıza göre güncelleyin
# ============================================================
ANKET_WORKBOOK = "Anket_Verileri.xlsx"   # Tüm dönemleri içeren workbook
                                          # (her sheet = bir dönem adı)
CIKTI_KLASOR   = "."                      # Çıktı klasörü (varsayılan: aynı dizin)


# ============================================================
# LİKERT DÖNÜŞÜM HARİTASI
# (her iki formatta da çalışır: sayı ve metin)
# ============================================================
LIKERT_METIN = {
    "kesinlikle katılmıyorum" : 1,
    "katılmıyorum"            : 2,
    "pek fazla katılmıyorum"  : 3,
    "katılıyorum"             : 4,
    "biraz katılıyorum"       : 5,
    "tamamen katılıyorum"     : 6,
}


def likert_donustur(deger):
    """Metin veya sayı Likert değerini 1-6 arası sayıya çevirir."""
    if pd.isna(deger):
        return None
    if isinstance(deger, (int, float)):
        v = int(deger)
        return v if 1 <= v <= 6 else None
    if isinstance(deger, str):
        temiz = deger.strip().lower()
        return LIKERT_METIN.get(temiz, None)
    return None


# ============================================================
# SORU GRUPLARI — ön eke göre kategori eşleşmesi
# ============================================================
GRUP_ONEKLER = {
    "Ders İçeriği"          : ("1_1", "3_1"),
    "Öğretim Elemanı"       : ("2_1", "4_1", "7_1", "8_1", "9_1", "10_1", "11_1"),
    "Ölçme Değerlendirme"   : ("12_1", "13_1"),
    "Yöntem"                : ("5_1", "6_1"),
    "Genel Memnuniyet"      : ("14_1", "15_1", "16_1"),
}

KATEGORI_SUTUNLAR = list(GRUP_ONEKLER.keys())

# ============================================================
# EXCEL BİÇİMLENDİRME SABİTLERİ
# ============================================================
RENK_BASLIK    = PatternFill("solid", fgColor="1F4E79")
RENK_SUTUN_HDR = PatternFill("solid", fgColor="2E75B6")
RENK_SATIR_ALT = PatternFill("solid", fgColor="D6E4F0")
RENK_BEYAZ     = PatternFill("solid", fgColor="FFFFFF")
RENK_TOPLAM    = PatternFill("solid", fgColor="FCE4D6")
ince           = Side(style="thin", color="B0B0B0")
KENAR          = Border(left=ince, right=ince, top=ince, bottom=ince)
ORTALA         = Alignment(horizontal="center", vertical="center", wrap_text=True)
SOLA           = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def h_yaz(cell, val, fill=RENK_BEYAZ, bold=False, fmt=None, align=ORTALA):
    cell.value      = val
    cell.fill       = fill
    cell.font       = Font(name="Arial", bold=bold, size=10,
                           color="FFFFFF" if fill == RENK_BASLIK else "000000")
    cell.alignment  = align
    cell.border     = KENAR
    if fmt:
        cell.number_format = fmt


# ============================================================
# KATEGORİ HESAPLAMA
# ============================================================
def kategori_belirle(satir):
    """Ders Kodu ilk rakamına göre sınıf, Ders Birim'e göre öğretim türü."""
    birim   = str(satir.get("Ders Birim", "") or "")
    uzaktan = "uzaktan" in birim.lower()
    kod     = str(satir.get("Ders Kodu", "") or "")
    rakam   = next((c for c in kod if c.isdigit()), "1")
    try:
        sinif = min(int(rakam), 2)
    except ValueError:
        sinif = 1
    ogretim = "Uzaktan Öğretim" if uzaktan else "Birinci Öğretim"
    return f"{sinif}. Sınıf - {ogretim}"


def hesapla_gruplama(df):
    """
    Ham anket df'inden ders bazında kategori ortalamaları hesaplar.
    Döner: gruplama DataFrame
    """
    # Soru sütunlarını bul
    soru_cols = [c for c in df.columns if str(c)[0].isdigit()]

    # Likert dönüşümü
    for s in soru_cols:
        df[s] = df[s].apply(likert_donustur)

    # Kategori etiketleri
    df["Kategori"] = df.apply(kategori_belirle, axis=1)

    # Her soru grubunun satır ortalaması
    for kat, onekler in GRUP_ONEKLER.items():
        ilgili = [c for c in soru_cols if any(c.startswith(o) for o in onekler)]
        df[f"__{kat}"] = df[ilgili].mean(axis=1) if ilgili else float("nan")

    # Ders Birim + Ders Kodu + Kategori bazında grupla
    gruplama = df.groupby(
        ["Ders Birim", "Ders Kodu", "Kategori"], dropna=False
    ).agg(
        **{"Katılımcı (N)"       : ("Ders Kodu", "count")},
        **{"Ders İçeriği"        : ("__Ders İçeriği",        "mean")},
        **{"Öğretim Elemanı"     : ("__Öğretim Elemanı",     "mean")},
        **{"Ölçme Değerlendirme" : ("__Ölçme Değerlendirme", "mean")},
        **{"Yöntem"              : ("__Yöntem",              "mean")},
        **{"Genel Memnuniyet"    : ("__Genel Memnuniyet",    "mean")},
    ).reset_index()

    return gruplama


# ============================================================
# ÖZET SAYFA YAZICI
# ============================================================
def ozet_sayfa_yaz(ws, df_bolum, baslik_str):
    SUTUNLAR = [
        "Ders Birim", "Ders Kodu", "Kategori", "Katılımcı (N)",
        "Ders İçeriği", "Öğretim Elemanı", "Ölçme Değerlendirme",
        "Yöntem", "Genel Memnuniyet"
    ]
    # Başlık
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value, c.fill = f"{baslik_str} – Ders Değerlendirme Özeti", RENK_BASLIK
    c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    c.alignment = ORTALA
    c.border    = KENAR
    ws.row_dimensions[1].height = 32

    # Sütun başlıkları
    for ci, ad in enumerate(SUTUNLAR, 1):
        h = ws.cell(2, ci)
        h.value, h.fill = ad, RENK_SUTUN_HDR
        h.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        h.alignment = ORTALA
        h.border    = KENAR
    ws.row_dimensions[2].height = 36

    # Veri satırları
    for ri, row in enumerate(df_bolum.itertuples(index=False), start=3):
        dolgu = RENK_SATIR_ALT if ri % 2 == 0 else RENK_BEYAZ
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            if ci in (5, 6, 7, 8, 9):
                h_yaz(cell, round(float(val), 2) if pd.notna(val) else "", dolgu, fmt="0.00")
            elif ci in (1, 3):
                h_yaz(cell, val, dolgu, align=SOLA)
            else:
                h_yaz(cell, val, dolgu)
        ws.row_dimensions[ri].height = 18

    son = 2 + len(df_bolum)

    # Bölüm ortalaması
    if len(df_bolum) > 0:
        or_r = son + 1
        ws.row_dimensions[or_r].height = 20
        h_yaz(ws.cell(or_r, 1), "BÖLÜM ORTALAMASI", RENK_TOPLAM, bold=True, align=SOLA)
        h_yaz(ws.cell(or_r, 2), "", RENK_TOPLAM)
        h_yaz(ws.cell(or_r, 3), "", RENK_TOPLAM)
        h_yaz(ws.cell(or_r, 4), int(df_bolum["Katılımcı (N)"].sum()), RENK_TOPLAM, bold=True)
        for ci, sc in enumerate(KATEGORI_SUTUNLAR, 5):
            h_yaz(ws.cell(or_r, ci),
                  round(df_bolum[sc].mean(), 2), RENK_TOPLAM, bold=True, fmt="0.00")

    # Sütun genişlikleri
    for ci, g in enumerate([30, 12, 28, 14, 13, 16, 20, 12, 18], 1):
        ws.column_dimensions[get_column_letter(ci)].width = g
    ws.freeze_panes = "A3"


# ============================================================
# GRAFİK OLUŞTURUCU  — Sınıf bazında ayrı
# ============================================================
SINIF_RENKLER = {
    "Birinci Öğretim" : "#2E75B6",
    "Uzaktan Öğretim" : "#ED7D31",
}

def grafik_olustur(gruplama, sinif_no, donem_adi):
    """
    sinif_no: 1 veya 2
    Döner: PNG dosya yolu (temp)
    """
    prefix  = f"{sinif_no}. Sınıf"
    df_plot = gruplama[gruplama["Kategori"].str.startswith(prefix)].copy()

    if df_plot.empty:
        return None

    fig, axes = plt.subplots(1, 2, figsize=(16, max(5, len(df_plot) * 0.45 + 2)),
                             sharey=False)
    fig.suptitle(f"{donem_adi}  |  {sinif_no}. Sınıf Ders Memnuniyet Ortalamaları",
                 fontsize=13, fontweight="bold", y=1.01)

    ogretim_turleri = ["Birinci Öğretim", "Uzaktan Öğretim"]

    for ax, tur in zip(axes, ogretim_turleri):
        df_tur = df_plot[df_plot["Kategori"].str.contains(tur)].copy()
        ax.set_title(tur, fontsize=11, fontweight="bold", pad=8)
        ax.set_xlim(1, 6)
        ax.axvline(x=3.5, color="gray", linestyle="--", linewidth=0.8, alpha=0.6)

        if df_tur.empty:
            ax.text(3.5, 0.5, "Veri yok", ha="center", va="center",
                    transform=ax.transAxes, fontsize=10, color="gray")
            ax.set_xlabel("Ortalama (1-6)")
            continue

        etketler = df_tur["Ders Kodu"].astype(str) + "\n" + df_tur["Ders Birim"].astype(str).str[:25]
        x = range(len(df_tur))

        bar_w = 0.14
        for i, (kat, renk) in enumerate([
            ("Ders İçeriği",        "#1F4E79"),
            ("Öğretim Elemanı",     "#2E75B6"),
            ("Ölçme Değerlendirme", "#70AD47"),
            ("Yöntem",              "#FFC000"),
            ("Genel Memnuniyet",    "#ED7D31"),
        ]):
            offsets = [xi + (i - 2) * bar_w for xi in x]
            vals    = df_tur[kat].fillna(0).tolist()
            ax.barh(offsets, vals, height=bar_w * 0.9,
                    label=kat, color=renk, alpha=0.88)
            for ox, v in zip(offsets, vals):
                if v > 0:
                    ax.text(v + 0.04, ox, f"{v:.2f}", va="center",
                            fontsize=6.5, color="#333333")

        ax.set_yticks(list(x))
        ax.set_yticklabels(etketler, fontsize=7.5)
        ax.set_xlabel("Ortalama (1-6)", fontsize=9)
        ax.legend(loc="lower right", fontsize=7, framealpha=0.7)
        ax.grid(axis="x", linestyle=":", alpha=0.5)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    plt.tight_layout()

    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    plt.savefig(tmp.name, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return tmp.name


def grafik_sayfasi_ekle(wb, png_yolu, sayfa_adi):
    """PNG'yi Excel sayfasına göm."""
    ws = wb.create_sheet(sayfa_adi)
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[1].height = 600
    img = XLImage(png_yolu)
    img.anchor = "A1"
    ws.add_image(img)


# ============================================================
# ANA FONKSİYON
# ============================================================
def isle_donem(df_ham, donem_adi, cikti_klasor="."):
    """
    Tek bir dönemin ham DataFrame'ini işler ve
    <donem_adi>_OZET.xlsx dosyasını kaydeder.
    """
    print(f"\n  ► {donem_adi} işleniyor ({len(df_ham)} satır)...")

    gruplama = hesapla_gruplama(df_ham.copy())

    # --- Yeni workbook ---
    wb = Workbook()
    wb.remove(wb.active)  # boş varsayılan sayfayı kaldır

    # --- Tüm dersler özet ---
    ws_tum = wb.create_sheet("Tum_Dersler_Ozet")
    ozet_sayfa_yaz(ws_tum, gruplama, "Tüm Dersler")

    # --- Bölüm sayfaları ---
    bolumler = sorted(gruplama["Ders Birim"].dropna().unique())
    for bolum in bolumler:
        df_b = gruplama[gruplama["Ders Birim"] == bolum].copy()
        guvli = (bolum.replace("/", "-").replace("\\", "-")
                 .replace("?", "").replace("*", "")
                 .replace("[", "").replace("]", "").replace(":", "-"))
        sayfa = f"Bolum_{guvli}"[:31]
        ozet_sayfa_yaz(wb.create_sheet(sayfa), df_b, bolum)

    # --- Grafik sayfaları (1. ve 2. sınıf ayrı) ---
    for sinif in [1, 2]:
        png = grafik_olustur(gruplama, sinif, donem_adi)
        if png:
            grafik_sayfasi_ekle(wb, png, f"Sinif{sinif}_Grafik")
            os.unlink(png)

    # --- Kaydet ---
    cikti = os.path.join(cikti_klasor, f"{donem_adi}_OZET.xlsx")
    wb.save(cikti)
    print(f"    ✔ Kaydedildi → {cikti}")
    print(f"      Bölüm sayısı : {len(bolumler)}")
    print(f"      Toplam ders  : {len(gruplama)}")
    return cikti


def isle_workbook(workbook_yolu, cikti_klasor="."):
    """
    Birden fazla dönem içeren workbook'u okur,
    her sheet için ayrı çıktı dosyası üretir.
    """
    print("=" * 60)
    print("  GÖREV 2 — Memnuniyet Ortalamaları & Özet Sayfaları")
    print("=" * 60)

    sheets = pd.read_excel(workbook_yolu, sheet_name=None)
    print(f"\nWorkbook: {workbook_yolu}")
    print(f"Bulunan dönemler: {list(sheets.keys())}")

    olusturulan = []
    for donem_adi, df in sheets.items():
        cikti = isle_donem(df, donem_adi, cikti_klasor)
        olusturulan.append(cikti)

    print(f"\n{'=' * 60}")
    print(f"✔ Toplam {len(olusturulan)} dönem işlendi.")
    for f in olusturulan:
        print(f"   • {f}")


if __name__ == "__main__":
    isle_workbook(ANKET_WORKBOOK, CIKTI_KLASOR)
