
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# DOSYA YOLLARI  →  kendi bilgisayarınıza göre değiştirin
# ============================================================
BASARI_DOSYA  = "2025-2026_Güz_basari_oranı.xlsx"
TBMYO_DOSYA   = "tbmyo_2025-2026_guz.xlsx"
MD_DOSYA      = "MD_Akademik.xlsx"


# ============================================================
# GÖREV 1: Ders Birim aktarımı
# ============================================================

def gorev1_ders_birim_aktar():
    """
    basari_oranı dosyasındaki 'Birim' bilgisini,
    Ders Kodu + Grup No eşleşmesiyle tbmyo dosyasının
    'Ders Birim' sütununa yazar.
    """

    # ----- 1. Dosyaları oku -----
    df_basari = pd.read_excel(BASARI_DOSYA)
    df_tbmyo  = pd.read_excel(TBMYO_DOSYA)

    # ----- 2. Eşleştirme anahtarı oluştur -----
    # Grup No float olarak gelebilir (1.0), int'e çevirerek temizliyoruz
    df_basari["anahtar"] = (
        df_basari["Ders Kodu"].astype(str).str.strip()
        + "_"
        + df_basari["Grup No"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    )

    df_tbmyo["anahtar"] = (
        df_tbmyo["Ders Kodu"].astype(str).str.strip()
        + "_"
        + df_tbmyo["Grup No"].astype(str).str.strip()
    )

    # ----- 3. Birim sözlüğü oluştur ve tbmyo'ya aktar -----
    birim_sozluk = df_basari.set_index("anahtar")["Birim"].to_dict()

    # Eşleşen satırlara Birim yaz; eşleşmeyenler mevcut değerini korur
    df_tbmyo["Ders Birim"] = (
        df_tbmyo["anahtar"]
        .map(birim_sozluk)
        .fillna(df_tbmyo["Ders Birim"])
    )

    # Yardımcı anahtar sütununu kaldır
    df_tbmyo.drop(columns=["anahtar"], inplace=True)

    # ----- 4. Güncellenmiş tbmyo dosyasını kaydet -----
    # Orijinal dosyayı bozmamak için yeni isimle kaydediyoruz
    cikti_dosya = "tbmyo_2025-2026_guz_GUNCELLENDI.xlsx"

    wb = load_workbook(TBMYO_DOSYA)
    ws = wb.active

    # Ders Birim sütununun Excel'deki indeksini bul
    baslik = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    birim_sutun = baslik.index("Ders Birim") + 1  # 1 tabanlı indeks

    # Değerleri satır satır yaz (2. satırdan başla, 1. satır başlık)
    for i, deger in enumerate(df_tbmyo["Ders Birim"], start=2):
        ws.cell(row=i, column=birim_sutun).value = deger

    wb.save(cikti_dosya)

    # ----- 5. Sonuç raporu -----
    doldurulan = df_tbmyo["Ders Birim"].notna().sum()
    print(f"[GÖREV 1] Tamamlandı → '{cikti_dosya}'")
    print(f"  Toplam satır  : {len(df_tbmyo)}")
    print(f"  Doldurulan    : {doldurulan}")
    print(f"  Eşleşmeyen    : {len(df_tbmyo) - doldurulan}")

    return df_tbmyo   # Görev 2'de kullanılacak


# ============================================================
# GÖREV 2: Bölümlere göre özet sayfaları
# ============================================================

# --- Likert ölçeği metinden sayıya dönüşüm tablosu ---
LIKERT_MAP = {
    "Kesinlikle katılmıyorum" : 1,
    "Pek fazla katılmıyorum"  : 2,
    "Katılmıyorum"            : 3,
    "Biraz katılıyorum"       : 4,
    "Katılıyorum"             : 5,
    "Tamamen katılıyorum"     : 6,
}

# --- Soru grupları (hangi soru hangi kategoride) ---
# Sütun adları "N_1 ..." formatındadır; ön ek numarasına göre gruplandırıyoruz
GRUP_ONEKLER = {
    "Ders İçeriği"          : ("1_1", "3_1"),
    "Öğretim Elemanı"       : ("2_1", "4_1", "7_1", "8_1", "9_1", "10_1", "11_1"),
    "Ölçme Değerlendirme"   : ("12_1", "13_1"),
    "Yöntem"                : ("5_1", "6_1"),
    "Genel Memnuniyet"      : ("14_1", "15_1", "16_1"),
}

# --- Sayfa biçimlendirme renkleri ---
RENK_BASLIK     = PatternFill("solid", fgColor="1F4E79")   # koyu mavi
RENK_SUTUN_HDR  = PatternFill("solid", fgColor="2E75B6")   # orta mavi
RENK_SATIR_ALT  = PatternFill("solid", fgColor="D6E4F0")   # açık mavi
RENK_BEYAZ      = PatternFill("solid", fgColor="FFFFFF")
RENK_TOPLAM     = PatternFill("solid", fgColor="FCE4D6")   # açık turuncu
ince_kenar      = Side(style="thin", color="B0B0B0")
KENAR           = Border(left=ince_kenar, right=ince_kenar,
                         top=ince_kenar, bottom=ince_kenar)
ORTALA          = Alignment(horizontal="center", vertical="center", wrap_text=True)
SOLA            = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def hucre_yaz(hucre, deger, dolgu=RENK_BEYAZ, kalin=False, format_kodu=None, hizalama=ORTALA):
    """Bir hücreye değer, stil ve biçim uygular."""
    hucre.value      = deger
    hucre.fill       = dolgu
    hucre.font       = Font(name="Arial", bold=kalin, size=10,
                            color="FFFFFF" if dolgu == RENK_BASLIK else "000000")
    hucre.alignment  = hizalama
    hucre.border     = KENAR
    if format_kodu:
        hucre.number_format = format_kodu


def ozet_sayfasi_yaz(ws, df_bolum, bolum_adi):
    """
    Verilen worksheet'e tek bir bölümün özet tablosunu yazar.
    Tum_Dersler_Ozet sayfasındaki formata uygun düzenlenir.
    """
    SUTUNLAR = [
        "Ders Birim", "Ders Kodu", "Kategori", "Katılımcı (N)",
        "Ders İçeriği", "Öğretim Elemanı", "Ölçme Değerlendirme",
        "Yöntem", "Genel Memnuniyet"
    ]

    # --- Birleşik başlık satırı ---
    ws.merge_cells("A1:I1")
    hucre = ws["A1"]
    hucre.value     = f"{bolum_adi} – Ders Değerlendirme Özeti"
    hucre.fill      = RENK_BASLIK
    hucre.font      = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    hucre.alignment = ORTALA
    hucre.border    = KENAR
    ws.row_dimensions[1].height = 32

    # --- Sütun başlıkları (2. satır) ---
    for ci, ad in enumerate(SUTUNLAR, 1):
        h = ws.cell(2, ci)
        h.value     = ad
        h.fill      = RENK_SUTUN_HDR
        h.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        h.alignment = ORTALA
        h.border    = KENAR
    ws.row_dimensions[2].height = 36

    # --- Veri satırları ---
    for ri, satir in enumerate(df_bolum.itertuples(index=False), start=3):
        dolgu = RENK_SATIR_ALT if ri % 2 == 0 else RENK_BEYAZ
        degerler = list(satir)
        for ci, val in enumerate(degerler, 1):
            h = ws.cell(ri, ci)
            if ci in (5, 6, 7, 8, 9):                # sayısal ortalama sütunları
                hucre_yaz(h, round(float(val), 2), dolgu, format_kodu="0.00")
            elif ci in (1, 3):                        # metin sütunları → sola hizalı
                hucre_yaz(h, val, dolgu, hizalama=SOLA)
            else:
                hucre_yaz(h, val, dolgu)
        ws.row_dimensions[ri].height = 18

    son_veri_satiri = 2 + len(df_bolum)

    # --- Bölüm ortalaması satırı ---
    if len(df_bolum) > 0:
        or_satir = son_veri_satiri + 1
        ws.row_dimensions[or_satir].height = 20

        hucre_yaz(ws.cell(or_satir, 1), "BÖLÜM ORTALAMASI",
                  RENK_TOPLAM, kalin=True, hizalama=SOLA)
        hucre_yaz(ws.cell(or_satir, 2), "", RENK_TOPLAM)
        hucre_yaz(ws.cell(or_satir, 3), "", RENK_TOPLAM)
        hucre_yaz(ws.cell(or_satir, 4),
                  int(df_bolum["Katılımcı (N)"].sum()), RENK_TOPLAM, kalin=True)

        say_sutunlar = ["Ders İçeriği", "Öğretim Elemanı",
                        "Ölçme Değerlendirme", "Yöntem", "Genel Memnuniyet"]
        for ci, sc in enumerate(say_sutunlar, 5):
            hucre_yaz(ws.cell(or_satir, ci),
                      round(df_bolum[sc].mean(), 2),
                      RENK_TOPLAM, kalin=True, format_kodu="0.00")

    # --- Sütun genişlikleri ---
    genislikler = [30, 12, 28, 14, 13, 16, 20, 12, 18]
    for ci, g in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(ci)].width = g

    ws.freeze_panes = "A3"   # ilk iki satırı dondur


def gorev2_ozet_sayfalari_olustur(df_tbmyo):
    """
    tbmyo dataframe'inden ders bazında ortalamalar hesaplar,
    MD_Akademik.xlsx dosyasına:
      - Tum_Dersler_Ozet  (yenilenir)
      - Her bölüm için Bolum_<BolumAdi>  (yeni sayfa)
    ekler.
    """

    # ----- 1. Likert metinlerini sayıya çevir -----
    soru_sutunlar = [c for c in df_tbmyo.columns if c[0].isdigit()]
    for s in soru_sutunlar:
        df_tbmyo[s] = df_tbmyo[s].map(LIKERT_MAP)

    # ----- 2. Her kategori için soru sütunlarını belirle -----
    kat_sutunlar = {}
    for kat_adi, onekler in GRUP_ONEKLER.items():
        kat_sutunlar[kat_adi] = [c for c in soru_sutunlar
                                  if any(c.startswith(on) for on in onekler)]

    # ----- 3. Kategori etiketini belirle (1./2. sınıf, Birinci/Uzaktan) -----
    def kategori_belirle(satir):
        birim = str(satir.get("Ders Birim", ""))
        uzaktan = "Uzaktan" in birim

        # Ders kodunun ilk rakamına göre sınıf (örn. MDY1001 → 1. sınıf)
        kod = str(satir.get("Ders Kodu", ""))
        rakamlar = "".join(filter(str.isdigit, kod))
        try:
            sinif = min(int(rakamlar[0]), 2) if rakamlar else 1
        except (ValueError, IndexError):
            sinif = 1

        ogretim = "Uzaktan Öğretim" if uzaktan else "Birinci Öğretim"
        return f"{sinif}. Sınıf - {ogretim}"

    df_tbmyo["Kategori"] = df_tbmyo.apply(kategori_belirle, axis=1)

    # ----- 4. Her soru grubunun ortalamasını hesapla -----
    for kat_adi, sutunlar in kat_sutunlar.items():
        df_tbmyo[f"__{kat_adi}"] = df_tbmyo[sutunlar].mean(axis=1)

    # ----- 5. Ders Birim + Ders Kodu + Kategori bazında grupla -----
    gruplama = df_tbmyo.groupby(
        ["Ders Birim", "Ders Kodu", "Kategori"]
    ).agg(
        **{"Katılımcı (N)"       : ("Ders Kodu", "count")},
        **{"Ders İçeriği"        : (f"__Ders İçeriği",        "mean")},
        **{"Öğretim Elemanı"     : (f"__Öğretim Elemanı",     "mean")},
        **{"Ölçme Değerlendirme" : (f"__Ölçme Değerlendirme", "mean")},
        **{"Yöntem"              : (f"__Yöntem",              "mean")},
        **{"Genel Memnuniyet"    : (f"__Genel Memnuniyet",    "mean")},
    ).reset_index()

    # ----- 6. MD_Akademik dosyasını aç -----
    wb = load_workbook(MD_DOSYA)

    # Önceki özet sayfalarını temizle (varsa)
    silinecek = [s for s in wb.sheetnames
                 if s == "Tum_Dersler_Ozet" or s.startswith("Bolum_")]
    for s in silinecek:
        del wb[s]

    # ----- 7. Tüm dersler özet sayfası -----
    ws_tum = wb.create_sheet("Tum_Dersler_Ozet")
    ozet_sayfasi_yaz(ws_tum, gruplama, "Tüm Dersler")

    # ----- 8. Her bölüm için ayrı sayfa -----
    bolumler = sorted(gruplama["Ders Birim"].dropna().unique())

    for bolum in bolumler:
        df_b = gruplama[gruplama["Ders Birim"] == bolum].copy()

        # Excel sayfa adı max 31 karakter, özel karakterler yasak
        guvli_ad = (bolum
                    .replace("/", "-").replace("\\", "-")
                    .replace("?", "").replace("*", "")
                    .replace("[", "").replace("]", "")
                    .replace(":", "-"))
        sayfa_adi = f"Bolum_{guvli_ad}"[:31]

        ws_b = wb.create_sheet(sayfa_adi)
        ozet_sayfasi_yaz(ws_b, df_b, bolum)

    # ----- 9. Kaydet -----
    cikti_dosya = "MD_Akademik_OZET.xlsx"
    wb.save(cikti_dosya)

    print(f"\n[GÖREV 2] Tamamlandı → '{cikti_dosya}'")
    print(f"  Oluşturulan bölüm sayfası: {len(bolumler)}")
    for b in bolumler:
        print(f"    • {b}")


# ============================================================
# ANA PROGRAM
# ============================================================

if __name__ == "__main__":
    print("=" * 55)
    print("  2025-2026 Güz Dönemi Ders Değerlendirme Analizi")
    print("=" * 55)

    # Görev 1: Birim aktarımı → güncellenmiş df döner
    df_tbmyo_guncellenmis = gorev1_ders_birim_aktar()

    # Görev 2: Özet sayfaları oluştur
    gorev2_ozet_sayfalari_olustur(df_tbmyo_guncellenmis)

    print("\n✔ Her iki görev de başarıyla tamamlandı.")
